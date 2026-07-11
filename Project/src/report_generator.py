import logging
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any

logger = logging.getLogger("ETLEngine")

class ReportGenerator:
    def __init__(self, reports_dir: Path):
        self.reports_dir = reports_dir

    def _apply_styles(self, ws: openpyxl.worksheet.worksheet.Worksheet, title: str):
        """
        Applies a professional corporate styling theme to a worksheet.
        """
        # Header styles
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark blue
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[1].height = 28
        
        # Style headers
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Alternating row fills
        alt_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid") # Light grey-blue
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
            if row % 2 == 0:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = alt_fill

    def generate_duplicate_report(self, duplicates: List[Dict[str, Any]]) -> Path:
        """
        Generates Duplicate_Report.xlsx
        """
        file_path = self.reports_dir / "Duplicate_Report.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Duplicate Records"

        headers = [
            "VRN", 
            "Job Card No", 
            "Invoice No", 
            "Source File", 
            "Source Row", 
            "Confidence Score", 
            "Unselected Reason"
        ]
        
        # Write headers
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)

        # Write data
        for row_idx, dup in enumerate(duplicates, 2):
            ws.cell(row=row_idx, column=1, value=dup.get("vrn"))
            ws.cell(row=row_idx, column=2, value=dup.get("job_card_no"))
            ws.cell(row=row_idx, column=3, value=dup.get("invoice_no"))
            ws.cell(row=row_idx, column=4, value=dup.get("src_file"))
            ws.cell(row=row_idx, column=5, value=dup.get("src_row"))
            ws.cell(row=row_idx, column=6, value=dup.get("confidence_score"))
            ws.cell(row=row_idx, column=7, value=dup.get("unselected_reason"))

        self._apply_styles(ws, "Duplicate Records")
        wb.save(file_path)
        logger.info(f"Generated Duplicate Report at '{file_path.name}' with {len(duplicates)} rows.")
        return file_path

    def generate_conflict_report(self, conflicts: List[Dict[str, Any]]) -> Path:
        """
        Generates Conflict_Report.xlsx
        """
        file_path = self.reports_dir / "Conflict_Report.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Odometer Conflicts"

        headers = [
            "VRN", 
            "Service Date", 
            "Conflict Odometer", 
            "Source File", 
            "Source Row", 
            "Confidence Score", 
            "Other Values Found"
        ]

        # Write headers
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)

        # Write data
        for row_idx, con in enumerate(conflicts, 2):
            ws.cell(row=row_idx, column=1, value=con.get("vrn"))
            ws.cell(row=row_idx, column=2, value=con.get("date"))
            ws.cell(row=row_idx, column=3, value=con.get("odometer"))
            ws.cell(row=row_idx, column=4, value=con.get("source_file"))
            ws.cell(row=row_idx, column=5, value=con.get("source_row"))
            ws.cell(row=row_idx, column=6, value=con.get("confidence_score"))
            ws.cell(row=row_idx, column=7, value=str(con.get("other_values_found", [])))

        self._apply_styles(ws, "Odometer Conflicts")
        wb.save(file_path)
        logger.info(f"Generated Conflict Report at '{file_path.name}' with {len(conflicts)} rows.")
        return file_path

    def generate_validation_report(self, anomalies: List[Dict[str, Any]]) -> Path:
        """
        Generates Validation_Report.xlsx
        """
        file_path = self.reports_dir / "Validation_Report.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Validation Anomalies"

        headers = [
            "VRN", 
            "Anomaly Type", 
            "Severity", 
            "Previous Date", 
            "Previous Odometer", 
            "Current Date", 
            "Current Odometer", 
            "Detailed Description"
        ]

        # Write headers
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)

        # Write data
        for row_idx, anom in enumerate(anomalies, 2):
            ws.cell(row=row_idx, column=1, value=anom.get("vrn"))
            ws.cell(row=row_idx, column=2, value=anom.get("type"))
            ws.cell(row=row_idx, column=3, value=anom.get("severity"))
            ws.cell(row=row_idx, column=4, value=anom.get("prev_date"))
            ws.cell(row=row_idx, column=5, value=anom.get("prev_odo"))
            ws.cell(row=row_idx, column=6, value=anom.get("curr_date"))
            ws.cell(row=row_idx, column=7, value=anom.get("curr_odo"))
            ws.cell(row=row_idx, column=8, value=anom.get("description"))

        self._apply_styles(ws, "Validation Anomalies")
        wb.save(file_path)
        logger.info(f"Generated Validation Report at '{file_path.name}' with {len(anomalies)} rows.")
        return file_path

    def generate_merge_log(self, logs: List[Dict[str, Any]]) -> Path:
        """
        Generates Merge_Log.xlsx
        """
        file_path = self.reports_dir / "Merge_Log.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orchestration Log"

        headers = [
            "Timestamp", 
            "Phase / Step", 
            "Action Description", 
            "Details / Metrics"
        ]

        # Write headers
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)

        # Write data
        for row_idx, l in enumerate(logs, 2):
            ws.cell(row=row_idx, column=1, value=l.get("timestamp"))
            ws.cell(row=row_idx, column=2, value=l.get("phase"))
            ws.cell(row=row_idx, column=3, value=l.get("action"))
            ws.cell(row=row_idx, column=4, value=l.get("details"))

        self._apply_styles(ws, "Orchestration Log")
        wb.save(file_path)
        logger.info(f"Generated Merge Log at '{file_path.name}' with {len(logs)} rows.")
        return file_path
