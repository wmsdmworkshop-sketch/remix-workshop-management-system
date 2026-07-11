"""
DWIP Workforce v1.1 — Report Generator
======================================
Generates styled Excel exception and audit reports.
Supports single-sheet and multi-sheet workbooks with:
- Auto-filter tables
- Frozen header rows
- Auto-fit column widths
- Conditional formatting (traffic-light for scores/percentages)
- Summary sheet with timestamp, DWIP version, ValidationRunID, DB path
- Severity color coding (CRITICAL/WARNING/INFO)
Uses openpyxl for professional formatting.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

from etl.src.core.context import DWIPContext
from etl.src.core.step import ETLStep
from etl.src.core.result_types import MergeResult, ReportPackage

logger = logging.getLogger("dwip.report_generator")

# ── Premium Styling Palette ──────────────────────────────────────
NAVY_HEADER_FILL = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

# Severity Fills
CRITICAL_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
WARNING_FILL  = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
INFO_FILL     = PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid")

CRITICAL_FONT = Font(name="Segoe UI", size=10, bold=True, color="900C3F")
WARNING_FONT  = Font(name="Segoe UI", size=10, bold=True, color="7D6608")
INFO_FONT     = Font(name="Segoe UI", size=10, bold=True, color="1B4F72")

DATA_FONT = Font(name="Segoe UI", size=10, color="333333")

# Traffic light fills
GREEN_FILL  = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
AMBER_FILL  = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
RED_FILL    = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
GREEN_FONT  = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
AMBER_FONT  = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
RED_FONT    = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

# Summary sheet styles
SUMMARY_TITLE_FONT = Font(name="Segoe UI", size=14, bold=True, color="1B365D")
SUMMARY_LABEL_FONT = Font(name="Segoe UI", size=11, bold=True, color="333333")
SUMMARY_VALUE_FONT = Font(name="Segoe UI", size=11, color="555555")

thin_border_side = Side(border_style="thin", color="E0E0E0")
DATA_BORDER = Border(
    left=thin_border_side, right=thin_border_side,
    top=thin_border_side, bottom=thin_border_side,
)

# Thresholds for traffic-light
GREEN_THRESHOLD = 95.0
AMBER_THRESHOLD = 80.0


class ReportGenerator(ETLStep):
    """
    Generates beautiful styled Excel audit sheets.
    """

    def __init__(self, ctx: DWIPContext) -> None:
        super().__init__(ctx)
        self.logger = ctx.logger

    def execute(self, data: MergeResult) -> ReportPackage:
        # Return ReportPackage with summary
        return ReportPackage(report_paths=[], summary_json={}, status="PASS" if data.success else "FAIL")

    @staticmethod
    def write_dataframe_report(
        df: pd.DataFrame,
        output_path: Path,
        sheet_name: str = "Data",
        run_id: str = "",
        db_path: str = "",
        add_summary: bool = True,
        score_columns: Optional[List[str]] = None,
        severity_column: Optional[str] = None,
    ) -> None:
        """
        Write a single DataFrame to a styled Excel report.
        Includes auto-filter, freeze, auto-width, and conditional formatting.

        Args:
            df: DataFrame to export
            output_path: Path for the Excel file
            sheet_name: Name of the data sheet
            run_id: ValidationRunID for the summary sheet
            db_path: Database path for the summary sheet
            add_summary: Whether to add a Summary sheet
            score_columns: Column names to apply traffic-light formatting to
            severity_column: Column name containing CRITICAL/WARNING/INFO values
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]

        # Write data
        ReportGenerator._write_df_to_sheet(ws, df, severity_column)

        # Apply auto-filter
        if len(df) > 0:
            ws.auto_filter.ref = ws.dimensions

        # Apply traffic-light formatting on score columns
        if score_columns:
            ReportGenerator._apply_traffic_light(ws, df, score_columns)

        # Add summary sheet
        if add_summary:
            ReportGenerator._add_summary_sheet(wb, run_id, db_path, len(df), sheet_name)

        wb.save(output_path)
        logger.info("Report saved: %s (%d rows)", output_path.name, len(df))

    @staticmethod
    def write_multi_sheet_report(
        sheets: Dict[str, pd.DataFrame],
        output_path: Path,
        run_id: str = "",
        db_path: str = "",
        score_columns_map: Optional[Dict[str, List[str]]] = None,
    ) -> None:
        """
        Write multiple DataFrames as separate sheets in a single workbook.

        Args:
            sheets: dict of {sheet_name: DataFrame}
            output_path: Path for the Excel file
            run_id: ValidationRunID
            db_path: Database path
            score_columns_map: optional dict of {sheet_name: [score_col_names]}
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        score_columns_map = score_columns_map or {}

        wb = Workbook()

        first = True
        total_rows = 0
        for sheet_name, df in sheets.items():
            if first:
                ws = wb.active
                ws.title = sheet_name[:31]
                first = False
            else:
                ws = wb.create_sheet(title=sheet_name[:31])

            ReportGenerator._write_df_to_sheet(ws, df)

            if len(df) > 0:
                ws.auto_filter.ref = ws.dimensions

            score_cols = score_columns_map.get(sheet_name)
            if score_cols:
                ReportGenerator._apply_traffic_light(ws, df, score_cols)

            total_rows += len(df)

        # Summary sheet
        ReportGenerator._add_summary_sheet(
            wb, run_id, db_path, total_rows,
            f"{len(sheets)} sheets"
        )

        wb.save(output_path)
        logger.info("Multi-sheet report saved: %s (%d sheets, %d total rows)",
                     output_path.name, len(sheets), total_rows)

    # ── Legacy API (backward-compatible) ─────────────────────────

    @staticmethod
    def generate_validation_report(failures: List[Dict[str, Any]], output_path: Path,
                                   run_id: str = "", db_path: str = "") -> None:
        """Generate Validation_Report.xlsx from validation failure dicts."""
        headers = [
            "Rule ID", "Severity", "VRN", "Job Card No",
            "Field Name", "Raw Value", "Description",
            "Source File", "Source Row"
        ]
        rows = []
        for fail in failures:
            rows.append([
                fail.get("rule_id", ""),
                fail.get("severity", ""),
                fail.get("vrn", ""),
                fail.get("job_card_no", ""),
                fail.get("field_name", ""),
                str(fail.get("raw_value", "") if fail.get("raw_value") is not None else ""),
                fail.get("description", ""),
                fail.get("source_file", ""),
                fail.get("source_row", "")
            ])
        df = pd.DataFrame(rows, columns=headers)
        ReportGenerator.write_dataframe_report(
            df, output_path, "Validation Failures",
            run_id=run_id, db_path=db_path,
            severity_column="Severity",
        )

    @staticmethod
    def generate_duplicate_report(duplicates: List[Dict[str, Any]], output_path: Path,
                                  run_id: str = "", db_path: str = "") -> None:
        """Generate Duplicate_Report.xlsx from duplicate records."""
        headers = [
            "Dataset Type", "VRN", "Job Card No", "Invoice No",
            "Source File", "Source Row", "Confidence Score", "Unselected Reason"
        ]
        rows = []
        for dup in duplicates:
            rows.append([
                dup.get("dataset_type", ""),
                dup.get("vrn", ""),
                dup.get("job_card_no", ""),
                dup.get("invoice_no", ""),
                dup.get("source_file", ""),
                dup.get("source_row", ""),
                dup.get("confidence_score", 0.0),
                dup.get("unselected_reason", "")
            ])
        df = pd.DataFrame(rows, columns=headers)
        ReportGenerator.write_dataframe_report(
            df, output_path, "Duplicates Log",
            run_id=run_id, db_path=db_path,
            score_columns=["Confidence Score"],
        )

    @staticmethod
    def generate_conflict_report(conflicts: List[Dict[str, Any]], output_path: Path,
                                 run_id: str = "", db_path: str = "") -> None:
        """Generate Conflict_Report.xlsx from conflict records."""
        headers = [
            "Conflict Type", "VRN", "Service Date", "Field Name",
            "Value A", "Value B", "Source File A", "Source Row A",
            "Source File B", "Source Row B", "Resolution"
        ]
        rows = []
        for conf in conflicts:
            rows.append([
                conf.get("conflict_type", ""),
                conf.get("vrn", ""),
                conf.get("service_date", ""),
                conf.get("field_name", ""),
                conf.get("value_a", ""),
                conf.get("value_b", ""),
                conf.get("source_file_a", ""),
                conf.get("source_row_a", ""),
                conf.get("source_file_b", ""),
                conf.get("source_row_b", ""),
                conf.get("resolution", "UNRESOLVED")
            ])
        df = pd.DataFrame(rows, columns=headers)
        ReportGenerator.write_dataframe_report(
            df, output_path, "Conflicts Log",
            run_id=run_id, db_path=db_path,
        )

    @staticmethod
    def generate_rejected_records_report(rejected: List[Dict[str, Any]], output_path: Path,
                                         run_id: str = "", db_path: str = "") -> None:
        """Generate Rejected_Records.xlsx."""
        headers = ["Reason", "Job Card No", "VRN", "Source File", "Source Row", "Raw Data JSON"]
        rows = []
        for r in rejected:
            rows.append([
                r.get("reason", ""),
                r.get("job_card_no", ""),
                r.get("vrn", ""),
                r.get("source_file", ""),
                r.get("source_row", ""),
                r.get("raw_data_json", "")
            ])
        df = pd.DataFrame(rows, columns=headers)
        ReportGenerator.write_dataframe_report(
            df, output_path, "Rejected Records",
            run_id=run_id, db_path=db_path,
        )

    @staticmethod
    def generate_merge_log_report(logs: List[Dict[str, Any]], output_path: Path,
                                  run_id: str = "", db_path: str = "") -> None:
        """Generate Merge_Log.xlsx."""
        headers = ["Phase", "Action", "Details", "Record Count", "Timestamp"]
        rows = []
        for log in logs:
            rows.append([
                log.get("phase", ""),
                log.get("action", ""),
                log.get("details", ""),
                log.get("record_count", ""),
                log.get("created_at", "")
            ])
        df = pd.DataFrame(rows, columns=headers)
        ReportGenerator.write_dataframe_report(
            df, output_path, "Execution Merge Log",
            run_id=run_id, db_path=db_path,
        )

    # ── Private helpers ──────────────────────────────────────────

    @staticmethod
    def _write_df_to_sheet(
        ws, df: pd.DataFrame,
        severity_column: Optional[str] = None,
    ) -> None:
        """Write a DataFrame to an openpyxl worksheet with styling."""
        # Header row
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = NAVY_HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 28

        # Data rows
        sev_col_idx = None
        if severity_column and severity_column in df.columns:
            sev_col_idx = list(df.columns).index(severity_column) + 1

        for r_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.border = DATA_BORDER
                cell.alignment = Alignment(vertical="center")

                # Severity colour coding
                if col_idx == sev_col_idx:
                    sev = str(value).upper()
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if sev == "CRITICAL":
                        cell.fill = CRITICAL_FILL
                        cell.font = CRITICAL_FONT
                    elif sev == "WARNING":
                        cell.fill = WARNING_FILL
                        cell.font = WARNING_FONT
                    elif sev == "INFO":
                        cell.fill = INFO_FILL
                        cell.font = INFO_FONT

            ws.row_dimensions[r_idx].height = 20

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 60)

    @staticmethod
    def _apply_traffic_light(
        ws, df: pd.DataFrame, score_columns: List[str],
    ) -> None:
        """Apply green/amber/red conditional formatting on score columns."""
        for col_name in score_columns:
            if col_name not in df.columns:
                continue
            col_idx = list(df.columns).index(col_name) + 1
            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}2:{col_letter}{len(df) + 1}"

            # Green >= 95
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(
                    operator="greaterThanOrEqual", formula=["95"],
                    fill=GREEN_FILL, font=GREEN_FONT,
                ),
            )
            # Amber >= 80
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(
                    operator="between", formula=["80", "94.99"],
                    fill=AMBER_FILL, font=AMBER_FONT,
                ),
            )
            # Red < 80
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(
                    operator="lessThan", formula=["80"],
                    fill=RED_FILL, font=RED_FONT,
                ),
            )

    @staticmethod
    def _add_summary_sheet(
        wb: Workbook,
        run_id: str,
        db_path: str,
        total_rows: int,
        data_description: str,
    ) -> None:
        """Add a Summary sheet with generation metadata."""
        ws = wb.create_sheet(title="Summary", index=0)

        rows = [
            ("DWIP Validation Report", ""),
            ("", ""),
            ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("DWIP Version", "1.1"),
            ("ETL Version", "2.5.0"),
            ("ValidationRunID", run_id or "N/A"),
            ("Database", db_path or "N/A"),
            ("Total Rows", total_rows),
            ("Data", data_description),
        ]

        for r_idx, (label, value) in enumerate(rows, start=1):
            ws.cell(row=r_idx, column=1, value=label).font = (
                SUMMARY_TITLE_FONT if r_idx == 1 else SUMMARY_LABEL_FONT
            )
            ws.cell(row=r_idx, column=2, value=value).font = SUMMARY_VALUE_FONT

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 50
